"""Export the study's metric tables to formatted Excel.

Two workbooks (same data, winners recomputed per version):
  results/metric_tables.xlsx          -- all four models
  results/metric_tables_noCatGP.xlsx  -- without the Categorical GP

Each workbook: one sheet per (problem, budget) config with the n_rep=10 table on top and the
n_rep=3 table below (rows = levels + Mean, column blocks = models x MAE/RRMSE/IS/Coverage),
plus a "summary" sheet organized like the LaTeX summary: merged Problem cell spanning its
budgets, merged n cell spanning its models, one model per row, column blocks n_rep=10 /
n_rep=3. Winner per comparison group (bold dark-green on light green): min for MAE/RRMSE/IS,
closest to 0.90 for Coverage.
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import problems as P, study, metrics as MET

LABELS = {"separate_gp": "Separate GPs", "categorical_kernel": "Categorical GP",
          "standard_LVGP": "Standard LVGP", "heter_LVGP": "H-LVGP"}
SETS = [("metric_tables.xlsx", ["separate_gp", "categorical_kernel", "standard_LVGP",
                                "heter_LVGP"]),
        ("metric_tables_noCatGP.xlsx", ["separate_gp", "standard_LVGP", "heter_LVGP"])]
METRICS = ["MAE", "RRMSE", "IS", "Coverage"]
PROBLEMS = [("branin_hetero", "Branin", [(None, 10), (25, 25), (40, 40)], "branin"),
            ("sixhump_camel", "Six-hump camel", [(None, 8), (20, 20), (32, 32)], "camel"),
            ("rastrigin_6d", "Rastrigin 6-D", [(20, 20), (None, 32)], "rastrigin")]
SEED, N_REPS = 1, [10, 3]

BOLD_GREEN = Font(bold=True, color="1F6B3B")
FILL_GREEN = PatternFill("solid", fgColor="EAF5EE")
HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILLS = {"Separate GPs": "3D6F9E", "Categorical GP": "8A5FA8",
             "Standard LVGP": "C07A3A", "H-LVGP": "2E8B6E"}
NREP_FILL = PatternFill("solid", fgColor="44546A")
THIN = Border(left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
              top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"))
CENTER = Alignment(horizontal="center", vertical="center")

_cache = {}


def table_of(prob, n_init, n_rep, key):
    ck = (prob, n_init, n_rep, key)
    if ck not in _cache:
        spec = P.get(prob)
        _cache[ck] = MET.metrics_table(
            spec, study.CachedPredictions(prob, key, SEED, n_rep, n_init), study.test_points(spec))
    return _cache[ck]


def best_model(vals, metric):
    if metric == "Coverage":
        return min(vals, key=lambda m: abs(vals[m] - 0.90))
    return min(vals, key=vals.get)


def put(ws, r, c, v, win=False, highlight=True):
    """Winner formatting: bold always; the green fill only when `highlight` (Mean rows)."""
    cell = ws.cell(r, c, round(v, 4))
    cell.number_format = "0.0000"; cell.alignment = CENTER; cell.border = THIN
    if win:
        cell.font = BOLD_GREEN if highlight else Font(bold=True)
        if highlight:
            cell.fill = FILL_GREEN


def config_sheet(ws, prob, keys, n_init, n_shown):
    models = [LABELS[k] for k in keys]
    spec = P.get(prob)
    ws.column_dimensions["A"].width = 12
    for col in range(2, 2 + len(models) * len(METRICS)):
        ws.column_dimensions[get_column_letter(col)].width = 10
    r0 = 1
    for n_rep in N_REPS:
        T = {LABELS[k]: table_of(prob, n_init, n_rep, k) for k in keys}
        levels = list(T[models[0]].index)
        t = ws.cell(r0, 1, f"{spec.name}  |  n_train = {n_shown} "
                           f"({n_shown // spec.n_levels}/level)  |  n_rep = {n_rep}  |  seed {SEED}")
        t.font = Font(bold=True, size=12)
        hdr, sub = r0 + 1, r0 + 2
        ws.cell(sub, 1, "Level").font = Font(bold=True)
        for mi, m in enumerate(models):
            c0 = 2 + mi * len(METRICS)
            ws.merge_cells(start_row=hdr, start_column=c0, end_row=hdr,
                           end_column=c0 + len(METRICS) - 1)
            h = ws.cell(hdr, c0, m); h.font = HDR_FONT; h.alignment = CENTER
            h.fill = PatternFill("solid", fgColor=HDR_FILLS[m])
            for j, met in enumerate(METRICS):
                s = ws.cell(sub, c0 + j, met); s.font = Font(bold=True); s.alignment = CENTER
        for li, lvl in enumerate(levels):
            r = sub + 1 + li
            ws.cell(r, 1, lvl).font = Font(bold=(lvl == "Mean"))
            for met in METRICS:
                vals = {m: float(T[m].loc[lvl, met]) for m in models}
                win = best_model(vals, met)
                for mi, m in enumerate(models):
                    put(ws, r, 2 + mi * len(METRICS) + METRICS.index(met), vals[m],
                        m == win, highlight=(lvl == "Mean"))
        r0 = sub + 1 + len(levels) + 1
    ws.freeze_panes = "B4"


def summary_sheet(ws, keys):
    models = [LABELS[k] for k in keys]
    widths = {"A": 16, "B": 6, "C": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for col in range(4, 4 + len(N_REPS) * len(METRICS)):
        ws.column_dimensions[get_column_letter(col)].width = 10
    hdr, sub = 1, 2
    for c, name in [(1, "Problem"), (2, "n"), (3, "Model")]:
        ws.merge_cells(start_row=hdr, start_column=c, end_row=sub, end_column=c)
        h = ws.cell(hdr, c, name); h.font = Font(bold=True); h.alignment = CENTER
    for ni, nr in enumerate(N_REPS):
        c0 = 4 + ni * len(METRICS)
        ws.merge_cells(start_row=hdr, start_column=c0, end_row=hdr,
                       end_column=c0 + len(METRICS) - 1)
        h = ws.cell(hdr, c0, f"n_rep = {nr}"); h.font = HDR_FONT; h.alignment = CENTER
        h.fill = NREP_FILL
        for j, met in enumerate(METRICS):
            s = ws.cell(sub, c0 + j, met); s.font = Font(bold=True); s.alignment = CENTER
    r = sub + 1
    for prob, disp, budgets, _ in PROBLEMS:
        p0 = r
        for n_init, n_shown in budgets:
            b0 = r
            per_rep = {nr: {LABELS[k]: table_of(prob, n_init, nr, k) for k in keys}
                       for nr in N_REPS}
            for m in models:
                ws.cell(r, 3, m).border = THIN
                for ni, nr in enumerate(N_REPS):
                    T = per_rep[nr]
                    for j, met in enumerate(METRICS):
                        vals = {mm: float(T[mm].loc["Mean", met]) for mm in models}
                        put(ws, r, 4 + ni * len(METRICS) + j, vals[m],
                            m == best_model(vals, met))
                r += 1
            ws.merge_cells(start_row=b0, start_column=2, end_row=r - 1, end_column=2)
            b = ws.cell(b0, 2, n_shown); b.font = Font(bold=True); b.alignment = CENTER
        ws.merge_cells(start_row=p0, start_column=1, end_row=r - 1, end_column=1)
        pcell = ws.cell(p0, 1, disp); pcell.font = Font(bold=True, size=11)
        pcell.alignment = CENTER
    for row in ws.iter_rows(min_row=sub + 1, max_row=r - 1, min_col=1, max_col=3):
        for c in row:
            c.border = THIN
    ws.freeze_panes = "D3"


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    for fname, keys in SETS:
        wb = Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("summary")
        summary_sheet(ws, keys)
        for prob, disp, budgets, short in PROBLEMS:
            for n_init, n_shown in budgets:
                config_sheet(wb.create_sheet(f"{short}_n{n_shown}"), prob, keys, n_init, n_shown)
        wb.save(os.path.join(here, fname))
        print("wrote", fname, flush=True)


if __name__ == "__main__":
    main()
