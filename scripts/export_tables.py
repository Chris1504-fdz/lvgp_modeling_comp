"""Export the study's metric tables to formatted Excel.

Two workbooks (same data, winners recomputed per version):
  results/metric_tables.xlsx          -- all four models
  results/metric_tables_noCatGP.xlsx  -- without the Categorical GP

Each workbook: one sheet per (problem, budget) config with the n_rep=10 table on top and the
n_rep=3 table below (rows = levels + Mean, column blocks = models x MAE/RRMSE/IS/Coverage),
plus a "summary" sheet organized like the LaTeX summary: merged Problem cell spanning its
budgets, merged n cell spanning its models, one model per row, column blocks n_rep=10 /
n_rep=3. Winner per comparison group (bold dark-green on light green): min for MAE/RRMSE/IS,
closest to 0.95 for Coverage.
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import problems as P, study, metrics as MET

LABELS = {"separate_gp": "Separate GPs", "categorical_kernel": "Categorical GP",
          "standard_LVGP": "Standard LVGP", "heter_LVGP": "H-LVGP"}
SETS = [("metric_tables.xlsx", ["separate_gp", "categorical_kernel", "standard_LVGP",
                                "heter_LVGP"]),
        ("metric_tables_noCatGP.xlsx", ["separate_gp", "standard_LVGP", "heter_LVGP"])]
METRICS = ["MAE", "RRMSE", "IS", "Coverage"]
PROBLEMS = [("branin_hetero", "Branin", [(25, 25), (40, 40)], "branin"),
            ("sixhump_camel", "Six-hump camel", [(25, 25), (40, 40)], "camel"),
            ("griewank_6d", "Griewank 6-D", [(48, 48), (64, 64)], "griewank")]
SEED, N_REPS = 1, [10]

BOLD_GREEN = Font(bold=True, color="1F6B3B")
FILL_GREEN = PatternFill("solid", fgColor="EAF5EE")
HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILLS = {"Separate GPs": "3D6F9E", "Categorical GP": "8A5FA8",
             "Standard LVGP": "C07A3A", "H-LVGP": "2E8B6E"}
NREP_FILL = PatternFill("solid", fgColor="44546A")
THIN = Border(left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
              top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"))
CENTER = Alignment(horizontal="center", vertical="center")

_cache = {}


def table_of(prob, n_init, n_rep, key):
    ck = (prob, n_init, n_rep, key)
    if ck not in _cache:
        spec = P.get(prob)
        _cache[ck] = MET.metrics_table(
            spec, study.CachedPredictions(prob, key, SEED, n_rep, n_init), study.test_points(spec))
    return _cache[ck]


def best_model(vals, metric):
    if metric == "Coverage":
        return min(vals, key=lambda m: abs(vals[m] - 0.95))
    return min(vals, key=vals.get)


def put(ws, r, c, v, win=False, highlight=True):
    """Winner formatting: bold always; the green fill only when `highlight` (Mean rows)."""
    cell = ws.cell(r, c, round(v, 4))
    cell.number_format = "0.0000"; cell.alignment = CENTER; cell.border = THIN
    if win:
        cell.font = BOLD_GREEN if highlight else Font(bold=True)
        if highlight:
            cell.fill = FILL_GREEN


def config_sheet(ws, prob, keys, n_init, n_shown):
    models = [LABELS[k] for k in keys]
    spec = P.get(prob)
    ws.column_dimensions["A"].width = 12
    for col in range(2, 2 + len(models) * len(METRICS)):
        ws.column_dimensions[get_column_letter(col)].width = 10
    r0 = 1
    for n_rep in N_REPS:
        T = {LABELS[k]: table_of(prob, n_init, n_rep, k) for k in keys}
        levels = list(T[models[0]].index)
        t = ws.cell(r0, 1, f"{spec.name}  |  n_train = {n_shown} "
                           f"({n_shown // spec.n_levels}/level)  |  n_rep = {n_rep}  |  seed {SEED}")
        t.font = Font(bold=True, size=12)
        hdr, sub = r0 + 1, r0 + 2
        ws.cell(sub, 1, "Level").font = Font(bold=True)
        for mi, m in enumerate(models):
            c0 = 2 + mi * len(METRICS)
            ws.merge_cells(start_row=hdr, start_column=c0, end_row=hdr,
                           end_column=c0 + len(METRICS) - 1)
            h = ws.cell(hdr, c0, m); h.font = HDR_FONT; h.alignment = CENTER
            h.fill = PatternFill("solid", fgColor=HDR_FILLS[m])
            for j, met in enumerate(METRICS):
                s = ws.cell(sub, c0 + j, met); s.font = Font(bold=True); s.alignment = CENTER
        for li, lvl in enumerate(levels):
            r = sub + 1 + li
            ws.cell(r, 1, lvl).font = Font(bold=(lvl == "Mean"))
            for met in METRICS:
                vals = {m: float(T[m].loc[lvl, met]) for m in models}
                win = best_model(vals, met)
                for mi, m in enumerate(models):
                    put(ws, r, 2 + mi * len(METRICS) + METRICS.index(met), vals[m],
                        m == win, highlight=(lvl == "Mean"))
        r0 = sub + 1 + len(levels) + 1
    ws.freeze_panes = "B4"


def summary_sheet(ws, keys):
    models = [LABELS[k] for k in keys]
    widths = {"A": 16, "B": 6, "C": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for col in range(4, 4 + len(N_REPS) * len(METRICS)):
        ws.column_dimensions[get_column_letter(col)].width = 10
    hdr, sub = 1, 2
    for c, name in [(1, "Problem"), (2, "n"), (3, "Model")]:
        ws.merge_cells(start_row=hdr, start_column=c, end_row=sub, end_column=c)
        h = ws.cell(hdr, c, name); h.font = Font(bold=True); h.alignment = CENTER
    for ni, nr in enumerate(N_REPS):
        c0 = 4 + ni * len(METRICS)
        ws.merge_cells(start_row=hdr, start_column=c0, end_row=hdr,
                       end_column=c0 + len(METRICS) - 1)
        h = ws.cell(hdr, c0, f"n_rep = {nr}"); h.font = HDR_FONT; h.alignment = CENTER
        h.fill = NREP_FILL
        for j, met in enumerate(METRICS):
            s = ws.cell(sub, c0 + j, met); s.font = Font(bold=True); s.alignment = CENTER
    r = sub + 1
    for prob, disp, budgets, _ in PROBLEMS:
        p0 = r
        for n_init, n_shown in budgets:
            b0 = r
            per_rep = {nr: {LABELS[k]: table_of(prob, n_init, nr, k) for k in keys}
                       for nr in N_REPS}
            for m in models:
                ws.cell(r, 3, m).border = THIN
                for ni, nr in enumerate(N_REPS):
                    T = per_rep[nr]
                    for j, met in enumerate(METRICS):
                        vals = {mm: float(T[mm].loc["Mean", met]) for mm in models}
                        put(ws, r, 4 + ni * len(METRICS) + j, vals[m],
                            m == best_model(vals, met))
                r += 1
            ws.merge_cells(start_row=b0, start_column=2, end_row=r - 1, end_column=2)
            b = ws.cell(b0, 2, n_shown); b.font = Font(bold=True); b.alignment = CENTER
        ws.merge_cells(start_row=p0, start_column=1, end_row=r - 1, end_column=1)
        pcell = ws.cell(p0, 1, disp); pcell.font = Font(bold=True, size=11)
        pcell.alignment = CENTER
    for row in ws.iter_rows(min_row=sub + 1, max_row=r - 1, min_col=1, max_col=3):
        for c in row:
            c.border = THIN
    ws.freeze_panes = "D3"


CONCLUSIONS = [
    ("Status", "Preliminary: seed 1, n_rep = 10. Problem versions: branin v2 (region-contrast "
     "noise), camel v4 (5 levels: two pairs + bridge), griewank_6d v1.3 (paired levels, rescaled "
     "noise, tamed domain). Multi-seed replication runs once the setup is frozen."),
    ("1. Cross-level sharing dominates", "Separate GPs trail on accuracy in every configuration "
     "(~2-4x the best MAE on the 1-D problems); honest wide intervals score well only on "
     "pure-calibration metrics."),
    ("2. Branin v2: H-LVGP most accurate", "Best MAE/RRMSE at both budgets (2.55/1.35 vs CatGP "
     "4.76/1.88), coverage 1.00, best IS+NLL at n=25. CatGP takes IS at n=40 (14.1 vs 30.6): "
     "conservative H-LVGP intervals pay a width penalty in the high-noise region."),
    ("3. Camel: falsified hypothesis -> synthesis", "v2.0 tight pairs (4 lvls): H-LVGP won "
     "n=32 (data pooling). v2.1 widened: CatGP regained. v3 hierarchy: prediction FALSIFIED - "
     "CatGP won bigger (0.097 vs 0.142). v4 (5 levels: two pairs + bridge): H-LVGP decisively "
     "retakes 8/level (n=40: MAE 0.0743 vs 0.1008, IS 0.765 vs 1.001, best NLL); CatGP keeps "
     "5/level. Synthesis: the latent pays with >=8 pts/level AND >=5 structured levels."),
    ("4. Standard LVGP intervals unusable", "Coverage 0.22-0.78 vs the 0.95 target everywhere "
     "(camel v3 n=32: 0.290, NLL 3351); worst miscalibration areas; NLL up to thousands. Mean "
     "often competitive - the deficiency is purely the noise-blind uncertainty."),
    ("5. Noise recovery: polynomial limits", "Branin's sigmoid step is outside both polynomial "
     "families (nMAE 0.25-0.33; pooled advantage vanishes). Under-determined pooled poly "
     "(36 coefs vs 20-32 pts) explodes out-of-sample (nMAE up to ~1300); over-determined it "
     "recovers the surface well (0.215 at griewank n=160)."),
    ("6. Griewank 6-D: fixed via domain, H-LVGP wins top budget", "Isolation experiments: "
     "budget ladder ruled out data quantity (RRMSE >= 1.09 at 10/20/40 pts/level); noise "
     "rescale 0.05 -> 0.02 ruled out data quality (unchanged). Cause: the 5-way cosine product "
     "oscillates unresolvably over [-5,5]^5. v1.3 domain [-3,3]^5 makes it learnable (RRMSE "
     "0.64-0.82 at n=160), where H-LVGP takes the full scorecard (MAE 0.0433, RRMSE 0.643, IS "
     "0.488, cov 0.935) and both LVGPs beat both GPs. At n=40-80 CatGP and H-LVGP trade wins."),
    ("Recommendations", "Well-separated levels & >=5 pts/level: H-LVGP. Moderately structured / "
     "hierarchical / near-duplicate levels: Categorical GP. Intervals used downstream: never "
     "standard LVGP alone. Signal at/below the resolvable limit: fix the problem first."),
]



def conclusions_sheet(ws):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    t = ws.cell(1, 1, "Preliminary conclusions (seed 1, n_rep = 10, 95% metrics)")
    t.font = Font(bold=True, size=13)
    r = 3
    for head, body in CONCLUSIONS:
        h = ws.cell(r, 1, head); h.font = Font(bold=True)
        h.alignment = Alignment(vertical="top", wrap_text=True)
        b = ws.cell(r, 2, body)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(28, 14 * (len(body) // 105 + 1))
        r += 2


def main():
    # workbooks are regenerable artifacts -> results/ (git-ignored); final shared tables go to results_v2
    here = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "results")
    for fname, keys in SETS:
        wb = Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("summary")
        summary_sheet(ws, keys)
        conclusions_sheet(wb.create_sheet("conclusions"))
        for prob, disp, budgets, short in PROBLEMS:
            for n_init, n_shown in budgets:
                config_sheet(wb.create_sheet(f"{short}_n{n_shown}"), prob, keys, n_init, n_shown)
        wb.save(os.path.join(here, fname))
        print("wrote", fname, flush=True)


if __name__ == "__main__":
    main()
